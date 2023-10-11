"""
    Microsoft Excel Writer (XLSX Format)

    Copyright: 2021-2022 (c) Sahana Software Foundation

    Permission is hereby granted, free of charge, to any person
    obtaining a copy of this software and associated documentation
    files (the "Software"), to deal in the Software without
    restriction, including without limitation the rights to use,
    copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following
    conditions:

    The above copyright notice and this permission notice shall be
    included in all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
    EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
    OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
    NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
    HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
    WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
    OTHER DEALINGS IN THE SOFTWARE.
"""

__all__ = ("XLSXWriter",
           "XLSXPivotTableWriter",
           )

import datetime
import re

from io import BytesIO

from gluon import HTTP, current
from gluon.contenttype import contenttype

from ..tools import get_crud_string, s3_get_foreign_key, s3_str, s3_strip_markup, s3_has_foreign_key

from .base import FormatWriter

ROWS_PER_SHEET = 1048576

# =============================================================================
class XLSXWriter(FormatWriter):
    """ XLSX Writer """

    # -------------------------------------------------------------------------
    @classmethod
    def encode(cls, resource, **attr):
        """
            Export data as a Microsoft Excel spreadsheet

            Args:
                resource: the data source
                          - a CRUDResource
                          - a dict like: {columns: [key, ...],
                                          headers: {key: label},
                                          types: {key: type},
                                          rows: [{key:value}],
                                          }

            Keyword Args:
                title: the main title of the report
                list_fields: fields to include in list views
                use_color: use background colors in cells (boolean, default False)
                even_odd: when using colors, render different background colors
                          for even/odd rows (boolean, default True)
                as_stream: return BytesIO rather than bytes
        """

        T = current.T
        request = current.request
        settings = current.deployment_settings

        # Import libraries
        try:
            from openpyxl import Workbook
            from openpyxl.cell import Cell
            from openpyxl.utils import get_column_letter
        except ImportError:
            error = T("Export failed: OpenPyXL library not installed on server")
            current.log.error(error)
            raise HTTP(503, body=error)

        # Get the attributes
        attr_get = attr.get

        title = attr_get("title")
        if title is None:
            title = current.T("Report")

        list_fields = attr_get("list_fields")
        if isinstance(resource, dict):
            # Pre-extracted data dict
            headers = resource.get("headers", {})
            lfields = resource.get("columns", list_fields)
            rows = resource.get("rows")
            column_types = resource.get("types")
            types = [column_types[col] for col in lfields]
        else:
            # Extract data from CRUDResource
            if not list_fields:
                list_fields = resource.list_fields()
            title, types, lfields, headers, rows = cls.extract(resource, list_fields)

        # Verify columns in items
        if len(rows) > 0 and len(lfields) > len(rows[0]):
            missing = [s for s in lfields if s not in rows[0]]
            msg = "XLSXWriter: field(s) missing from data rows (%s)" % ", ".join(missing)
            current.log.error(msg)

        # Create the workbook
        wb = Workbook(iso_dates=True)

        # Add named styles
        use_color = attr_get("use_color", False)
        even_odd = attr_get("even_odd", True)
        cls.add_styles(wb, use_color=use_color, even_odd=even_odd)

        # Determine title row length and batch size
        title_row = settings.get_xls_title_row()
        if title_row:
            title_row_length = title_row(None) if callable(title_row) else 2
        else:
            title_row_length = 0
        batch_size = ROWS_PER_SHEET - title_row_length - 1

        # Generate columns labels
        labels = []
        for selector in lfields:
            label = headers[selector]
            if label in ("Id", "Sort"):
                continue
            labels.append(s3_str(label))

        # Add the work sheets

        # Characters /\?*[] not allowed in sheet names
        sheet_name = " ".join(re.sub(r"[\\\/\?\*\[\]:]", " ", s3_str(title)).split())

        batch, remaining = rows[:batch_size], rows[batch_size:]
        sheet_number = 0
        while batch:

            # Create work sheet
            sheet_number += 1
            ws_title = "%s-%s" % (sheet_name[:28], sheet_number)
            if sheet_number == 1:
                ws = wb.active
                ws.title = sheet_name[:31] if not remaining else ws_title
            else:
                ws = wb.create_sheet(title=ws_title)

            # Count columns and initialize column width
            num_columns = len(labels)
            column_widths = [len(label) for label in labels]

            # Add title row(s)
            if callable(title_row):
                title_row(ws)
            elif title_row:
                # First row: title
                top = Cell(ws, value=s3_str(title))
                top.style = "large_header"
                ws.append([top])

                # Second row: export date/time
                now = current.calendar.format_datetime(current.request.now, local=True)
                sub = Cell(ws, value="%s: %s" % (T("Date Exported"), now))
                sub.style = "header"
                ws.append([sub])

                # Adjust row dimension for large header
                row = ws.row_dimensions[1]
                row.height = 20

                # Merge columns
                if num_columns > 0:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)

            # Add column labels
            label_row = []
            for l in labels:
                cell = Cell(ws, value=l)
                cell.style = "label"
                label_row.append(cell)
            ws.append(label_row)

            # Add the data
            cls.write_rows(ws, batch, lfields, types, column_widths)

            # Adjust column widths
            for i in range(1, num_columns + 1):
                ws.column_dimensions[get_column_letter(i)].width = column_widths[i-1] * 1.23

            # Freeze title and column labels (=scroll only data rows)
            ws.freeze_panes = "A%d" % (title_row_length + 2)

            batch, remaining = remaining[:batch_size], remaining[batch_size:]

        # Save workbook
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            output = tmp.read()

        if not attr_get("as_stream", False):
            # Set response headers
            filename = "%s_%s.xlsx" % (request.env.server_name, title)
            disposition = "attachment; filename=\"%s\"" % filename
            response = current.response
            response.headers["Content-Type"] = contenttype(".xlsx")
            response.headers["Content-disposition"] = disposition
        else:
            # Convert to BytesIO
            output = BytesIO(output)

        return output

    # -------------------------------------------------------------------------
    @classmethod
    def write_rows(cls, ws, batch, lfields, types, column_widths):
        """
            Write the data rows

            Args:
                ws: the worksheet
                batch: the rows batch
                lfields: the column selectors
                types: the column types
                column_widths: mutable array of column widths
        """

        settings = current.deployment_settings

        # Date/Time formats from L10N settings
        date_format = settings.get_L10n_date_format()
        date_format_str = str(date_format)

        dtformats = {"date": dt_format_translate(date_format),
                     "time": dt_format_translate(settings.get_L10n_time_format()),
                     "datetime": dt_format_translate(settings.get_L10n_datetime_format()),
                     }

        from openpyxl.cell import Cell

        for i, row in enumerate(batch):
            outrow = []
            col_idx = 0
            for j, selector in enumerate(lfields):
                ftype = types[j]
                num_format = None

                if ftype in ("id", "sort"):
                    continue

                try:
                    value = s3_strip_markup(s3_str(row[selector]))
                except (KeyError, AttributeError):
                    value = ""
                width = len(value)
                if width > column_widths[col_idx]:
                    column_widths[col_idx] = width

                if ftype == "integer":
                    try:
                        value = to_int(value)
                    except ValueError:
                        pass
                    else:
                        num_format = "0"

                elif ftype == "double":
                    try:
                        value = to_float(value)
                    except ValueError:
                        pass
                    else:
                        num_format = "0.00"

                elif ftype in ("date", "datetime", "time"):
                    try:
                        value = datetime.datetime.strptime(value, date_format_str)
                    except (ValueError, TypeError):
                        pass
                    else:
                        num_format = dtformats[ftype]

                cell = Cell(ws, value=value)
                cell.style = "odd" if i % 2 else "even"
                if num_format:
                    cell.number_format = num_format
                outrow.append(cell)
                col_idx += 1
            ws.append(outrow)

    # -------------------------------------------------------------------------
    @classmethod
    def extract(cls, resource, list_fields):
        """
            Extract the rows from the resource

            Args:
                resource: the resource
                list_fields: fields to include in list views
        """

        title = get_crud_string(resource.tablename, "title_list")

        get_vars = dict(current.request.vars)
        get_vars["iColumns"] = len(list_fields)
        query, orderby, left = resource.datatable_filter(list_fields,
                                                         get_vars,
                                                         )
        resource.add_filter(query)

        if orderby is None:
            orderby = resource.get_config("orderby")

        # Hierarchical FK Expansion:
        # setting = {field_selector: [LevelLabel, LevelLabel, ...]}
        expand_hierarchy = resource.get_config("xls_expand_hierarchy")

        data = resource.select(list_fields,
                               left = left,
                               limit = None,
                               count = True,
                               getids = True,
                               orderby = orderby,
                               represent = True,
                               show_links = False,
                               raw_data = True if expand_hierarchy else False,
                               )

        rfields = data.rfields
        rows = data.rows

        types = []
        lfields = []
        heading = {}
        for rfield in rfields:
            if rfield.show:
                if expand_hierarchy:
                    levels = expand_hierarchy.get(rfield.selector)
                else:
                    levels = None
                if levels:
                    num_levels = len(levels)
                    colnames = cls.expand_hierarchy(rfield, num_levels, rows)
                    lfields.extend(colnames)
                    types.extend(["string"] * num_levels)
                    T = current.T
                    for i, colname in enumerate(colnames):
                        heading[colname] = T(levels[i])
                else:
                    lfields.append(rfield.colname)
                    heading[rfield.colname] = rfield.label or \
                                rfield.field.name.capitalize().replace("_", " ")
                    if rfield.ftype == "virtual":
                        types.append("string")
                    else:
                        types.append(rfield.ftype)

        return (title, types, lfields, heading, rows)

    # -------------------------------------------------------------------------
    @staticmethod
    def expand_hierarchy(rfield, num_levels, rows):
        """
            Expand a hierarchical foreign key column into one column
            per hierarchy level

            Args:
                rfield: the column (S3ResourceField)
                num_levels: the number of levels (from root)
                rows: the Rows from ResourceData

            Returns:
                list of keys (column names) for the inserted columns
        """

        field = rfield.field
        if not field or rfield.ftype[:9] != "reference":
            return []

        # Get the look-up table
        ktablename = s3_get_foreign_key(field, m2m=False)[0]
        if not ktablename:
            return []

        colname = rfield.colname
        represent = field.represent

        # Get the hierarchy
        from ..tools import S3Hierarchy
        h = S3Hierarchy(ktablename)
        if not h.config:
            return []

        # Collect the values from rows
        values = set()
        for row in rows:
            value = row["_row"][colname]
            if type(value) is list:
                value = value[0]
            values.add(value)

        # Generate the expanded values
        expanded = h.repr_expand(values,
                                 levels = num_levels,
                                 represent = represent,
                                 )

        # ...and add them into the rows
        colnames = ["%s__%s" % (colname, l) for l in range(num_levels)]
        for row in rows:
            value = row["_row"][colname]
            if type(value) is list:
                value = value[0]
            hcols = expanded.get(value)
            for level in range(num_levels):
                row[colnames[level]] = hcols[level] if hcols else None

        return colnames

    # -------------------------------------------------------------------------
    @staticmethod
    def add_styles(wb, use_color=False, even_odd=True):
        """
            Add custom styles to a workbook

            Args:
                wb: the workbook
                use_color: use color when styling cells
                even_odd: use different background color for even/odd rows
        """

        from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side

        side = Side(border_style="hair", color='FF000000')
        border = Border(top=side, right=side, bottom=side, left=side)

        font_normal = Font(name="Arial", size=10)
        font_bold = Font(name="Arial", size=10, bold=True)
        font_large = Font(name="Arial", size=14, bold=True)

        style = NamedStyle(name="normal")
        style.font = font_normal
        wb.add_named_style(style)

        style = NamedStyle(name="even")
        style.font = font_normal
        if use_color and even_odd:
            style.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="odd")
        style.font = font_normal
        if use_color and even_odd:
            style.fill = PatternFill(start_color="E7E7E7", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="label")
        style.font = font_bold
        if use_color:
            style.fill = PatternFill(start_color="BBDCED", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="large_header")
        style.font = font_large
        if use_color:
            style.fill = PatternFill(start_color="BDBDEC", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="header")
        style.font = font_bold
        if use_color:
            style.fill = PatternFill(start_color="BDBDEC", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="subheader")
        style.font = font_bold
        if use_color:
            style.fill = PatternFill(start_color="CCCCFF", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="subtotals")
        style.font = font_bold
        if use_color:
            style.fill = PatternFill(start_color="E6E6E6", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

        style = NamedStyle(name="totals")
        style.font = font_bold
        if use_color:
            style.fill = PatternFill(start_color="C6C6C6", fill_type="solid")
            style.border = border
        wb.add_named_style(style)

# =============================================================================
class XLSXPivotTableWriter:
    """
        XLSX writer for PivotTables

        TODO Support multiple layers (=write multiple sheets)
    """

    def __init__(self, pt):
        """
            Args:
                pt: the S3PivotTable to encode
        """

        self.pt = pt

        # Initialize properties
        self._styles = None
        self._formats = None

        self.lookup = {}
        self.valuemap = {}

    # -------------------------------------------------------------------------
    def encode(self, title):
        """
            Convert this pivot table into an XLS file

            Args:
                title: the title of the report

            Returns:
                the XLS workbook
        """

        T = current.T

        try:
            from openpyxl import Workbook
        except ImportError:
            error = T("XLSXPivotTableWriter: export failed, OpenPyXL library not installed on server")
            current.log.error(error)
            raise HTTP(503, body=error)

        TOTAL = s3_str(T("Total")).upper()

        pt = self.pt

        # Get report options
        report_options = pt.resource.get_config("report_options", {})

        # Report dimensions
        fact = pt.facts[0]
        layer = fact.layer
        rows_dim = pt.rows
        cols_dim = pt.cols

        numrows = pt.numrows
        numcols = pt.numcols

        # Resource fields for dimensions
        rfields = pt.rfields
        fact_rfield = rfields[fact.selector]
        rows_rfield = rfields[rows_dim] if rows_dim else None
        cols_rfield = rfields[cols_dim] if cols_dim else None

        # Dimension labels
        get_label = fact._get_field_label
        if rows_dim:
            rows_label = s3_str(get_label(rows_rfield, report_options.get("rows")))
        else:
            rows_label = ""
        if cols_dim:
            cols_label = s3_str(get_label(cols_rfield, report_options.get("cols")))
        else:
            cols_label = ""
        fact_label = s3_str(fact.get_label(fact_rfield, report_options.get("fact")))

        # Index of the column for row totals
        total_column = (numcols + 1) if cols_dim else 1

        # Sort+represent rows and columns
        rows, cols = self.sortrepr()

        # Create workbook and add styles
        book = Workbook(iso_dates=True)
        for style in self.styles.values():
            book.add_named_style(style)

        sheet = book.active
        write = self.write

        # Write header
        title_row = current.deployment_settings.get_xls_title_row()
        if callable(title_row):
            # Custom header (returns number of header rows)
            title_length = title_row(sheet)

        elif title_row:
            # Default header
            title_length = 2

            # Report title
            write(sheet, 0, 0, s3_str(title),
                  colspan = numcols + 2,
                  style = "title",
                  )

            # Current date/time (in local timezone)
            from ..tools import S3DateTime
            dt = S3DateTime.to_local(current.request.utcnow)
            write(sheet, 1, 0, dt, style = "subheader", numfmt = "datetime")

        else:
            # No header
            title_length = -1

        rowindex = title_length + 1

        # Fact label
        if rows_dim and cols_dim:
            write(sheet, rowindex, 0, fact_label, style="fact_label")

        # Columns axis title
        if cols_dim:
            write(sheet, rowindex, 1, cols_label,
                  colspan = numcols,
                  style = "axis_title",
                  )
            rowindex += 1

        # Row axis title
        write(sheet, rowindex, 0, rows_label, style="axis_title")

        # Column labels
        if cols_dim:
            for idx, col in enumerate(cols):
                write(sheet, rowindex, idx + 1, col[2]["text"],
                      style = "col_label",
                      )
            total_label = TOTAL
        else:
            # Use fact title as row total label if there is no column axis
            total_label = fact_label

        # Row totals label
        write(sheet, rowindex, total_column, total_label, style="total_right")

        # Determine the number format for cell values
        numfmt = self.number_format()
        totfmt = "integer" if fact.method in ("count", "list") else numfmt

        # Choose cell value style according to number format
        fact_style = "numeric" if numfmt else None

        # Get fact representation method
        if fact.method == "list":
            listrepr = self.listrepr
            fact_repr = pt._represents([layer])[fact.selector]
            fk = self.is_foreign_key(fact)
        else:
            listrepr = fact_repr = None
            fk = False

        # Write data rows (if any)
        rowindex += 1
        if rows_dim:
            icell = pt.cell
            for i in range(numrows):

                row = rows[i]

                # Row-label
                write(sheet, rowindex + i, 0, row[2]["text"],
                      style = "row_label",
                      )

                # Cell column values (if any)
                if cols_dim:
                    for j in range(numcols):
                        cell = icell[row[0]][cols[j][0]]
                        if listrepr:
                            value = listrepr(cell, fact_rfield, fact_repr, fk=fk)
                        else:
                            value = cell[layer]
                        write(sheet, rowindex + i, j + 1, value,
                              numfmt = numfmt,
                              style = fact_style,
                              )

                # Row-total
                write(sheet, rowindex + i, total_column, row[1],
                      style = "total",
                      numfmt = totfmt,
                      )

            rowindex += numrows
            total_label = TOTAL
        else:
            # Use fact label as column totals label if
            # there is no row dimension
            total_label = fact_label

        # Column totals label
        write(sheet, rowindex, 0, total_label, style="total_left")

        # Column totals
        if cols_dim:
            for i in range(numcols):
                write(sheet, rowindex, i + 1, cols[i][1],
                      style = "total",
                      numfmt = totfmt,
                      )

        # Grand total
        total = pt.totals[layer]
        write(sheet, rowindex, total_column, total,
              style = "grand_total",
              numfmt = totfmt,
              )

        # Save workbook to temp file and return the contents
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile() as tmp:
            book.save(tmp.name)
            tmp.seek(0)
            output = tmp.read()

        return output

    # -------------------------------------------------------------------------
    def write(self,
              sheet,
              rowindex,
              colindex,
              value,
              style=None,
              numfmt=None,
              rowspan=None,
              colspan=None,
              adjust=True
              ):
        """
            Write a value to a spreadsheet cell

            Args:
                sheet: the work sheet
                rowindex: the row index of the cell (0-based)
                colindex: the column index of the cell (0-based)
                value: the value to write
                style: a style name (see styles property)
                numfmt: a number format name (see formats property)
                rowspan: number of rows to merge
                colspan: number of columns to merge
                adjust: True to adjust column width and row height,
                        False to suppress automatic adjustment
        """

        rowindex += 1
        colindex += 1

        if type(value) is list:
            labels = [s3_str(v) for v in value]
            contents = "\n".join(labels)
        else:
            labels = [s3_str(value)]
            contents = value

        cell = sheet.cell(rowindex, colindex, value=contents)
        cell.style = style if style else "default"
        if numfmt:
            cell.number_format = self.formats.get(numfmt, "")

        # Apply rowspan and colspan
        rowspan = 0 if not rowspan or rowspan < 1 else rowspan - 1
        colspan = 0 if not colspan or colspan < 1 else colspan - 1
        if rowspan > 0 or colspan > 0:
            sheet.merge_cells(start_row = rowindex,
                              start_column = colindex,
                              end_row = rowindex + rowspan,
                              end_column = colindex + colspan,
                              )

        # Adjust column width and row height
        # NB approximations, no exact science (not possible except by
        #    enforcing a particular fixed-width font, which we don't
        #    want), so manual adjustments after export may still be
        #    necessary. Better solutions welcome!
        style = self.styles.get(style)
        if style and adjust:

            from openpyxl.utils import get_column_letter

            fontsize = float(style.font.sz)

            row = sheet.row_dimensions[rowindex]
            col = sheet.column_dimensions[get_column_letter(colindex)]

            # Adjust column width
            if not colspan:
                if labels:
                    width = min(max(len(l) for l in labels), 28) * 1.23
                else:
                    width = 0
                if width > col.width:
                    col.width = float(width)

            # Adjust row height
            if not rowspan:

                lineheight = 1.2 if style.font.bold else 1.0

                import math
                numlines = 0
                width = col.width * (colspan + 1)
                for label in labels:
                    numlines += math.ceil(len(label) * 1.23 * (fontsize / 10) / width)

                if numlines > 1:
                    lines = min(numlines, 10)
                    height = int((lines + 0.5 / lineheight) * fontsize * lineheight)
                else:
                    height = int(fontsize * lineheight)

                if not row.height or height > row.height:
                    row.height = height

    # -------------------------------------------------------------------------
    @property
    def styles(self):
        """
            Style definitions for pivot tables (lazy property)

            Returns:
                dict of named styles
        """

        styles = self._styles
        if styles is None:

            from openpyxl.styles import NamedStyle, Font, Alignment

            align = lambda v, h: Alignment(horizontal=h, vertical=v, wrapText=True)

            center = align("center", "center")
            centerleft = align("center", "left")
            #bottomcentered = align("bottom", "center")
            bottomleft = align("bottom", "left")
            bottomright = align("bottom", "right")
            topleft = align("top", "left")
            topright = align("top", "right")

            styles = {}

            def add_style(name, fontsize=10, bold=False, italic=False, align=None):

                style = NamedStyle(name)
                style.font = Font(name="Arial", size=fontsize, bold=bold, italic=italic)
                style.alignment = align

                styles[name] = style

            add_style("default", align=topleft)
            add_style("numeric", align=bottomright)
            add_style("title", fontsize=14, bold=True, align=bottomleft)
            add_style("subheader", fontsize=8, italic=True, align=bottomleft)
            add_style("row_label", bold=True, align=centerleft)
            add_style("col_label", bold=True, align=center)
            add_style("fact_label", fontsize=13, bold=True, align=center)
            add_style("axis_title", fontsize=11, bold=True, align=center)
            add_style("total", fontsize=11, bold=True, italic=True, align=topright)
            add_style("total_left", fontsize=11, bold=True, italic=True, align=topleft)
            add_style("total_right", fontsize=11, bold=True, italic=True, align=center)
            add_style("grand_total", fontsize=12, bold=True, italic=True, align=topright)

            self._styles = styles

        return styles

    # -------------------------------------------------------------------------
    @property
    def formats(self):
        """
            Number formats for pivot tables (lazy property)

            Returns:
                dict of format strings
        """

        formats = self._formats
        if formats is None:

            # Date/Time formats from L10N deployment settings
            settings = current.deployment_settings

            date_format = dt_format_translate(settings.get_L10n_date_format())
            datetime_format = dt_format_translate(settings.get_L10n_datetime_format())
            time_format = dt_format_translate(settings.get_L10n_time_format())

            formats = {
                "date": date_format,
                "datetime": datetime_format,
                "time": time_format,
                "integer": "0",
                "double": "0.00"
            }

            self._formats = formats

        return formats

    # -------------------------------------------------------------------------
    def is_foreign_key(self, fact):
        """
            Determines whether a fact references a foreign key

            Args:
                pt: the PivotTable
                fact: the PivotTableFact

            Returns:
                boolean
        """

        try:
            rfield = self.pt.rfields[fact.selector]
        except (AttributeError, KeyError):
            return False

        field = rfield.field

        return field is not None and s3_has_foreign_key(field)

    # -------------------------------------------------------------------------
    def number_format(self):
        """
            Determine the number format for this pivot table

            Returns:
                the number format key (see formats property)
        """

        numfmt = None

        pt = self.pt

        fact = pt.facts[0]
        rfield = pt.rfields[fact.selector]

        ftype = rfield.ftype

        if fact.method == "count":
            numfmt = "integer"

        elif ftype == "integer":
            if fact.method == "avg":
                # Average value of ints is a float
                numfmt = "double"
            else:
                numfmt = "integer"

        elif ftype in ("date", "datetime", "time", "double"):
            numfmt = ftype

        elif ftype == "virtual":
            # Probe the first value
            value = pt.cell[0][0][fact.layer]
            if isinstance(value, int):
                numfmt = "integer"
            elif isinstance(value, float):
                numfmt = "double"
            else:
                if isinstance(value, datetime.datetime):
                    numfmt = "datetime"
                elif isinstance(value, datetime.date):
                    numfmt = "date"
                elif isinstance(value, datetime.time):
                    numfmt = "time"

        return numfmt

    # -------------------------------------------------------------------------
    def sortrepr(self):
        """
            Sort and represent pivot table axes

            Returns:
                tuple (rows, cols), each a list of tuples:
                    (index,               ...the index of the row/column in
                                             the original cell array
                     total,               ...total value of the row/column
                     {value: axis_value,  ...group value of the row/column
                      text: axis_repr,    ...representation of the group value
                      },
                     )
        """

        pt = self.pt

        rfields = pt.rfields
        layer = pt.facts[0].layer

        # Sort rows
        rows_dim = pt.rows
        rows_rfield = rfields[rows_dim] if rows_dim else None
        row_repr = pt._represent_method(rows_dim)
        irows = pt.row
        rows = []
        for i in range(pt.numrows):
            irow = irows[i]
            header = {"value": irow.value,
                      "text": irow.text if "text" in irow
                                        else row_repr(irow.value),
                      }
            rows.append((i, irow[layer], header))
        pt._sortdim(rows, rows_rfield, index=2)

        # Sort columns
        cols_dim = pt.cols
        cols_rfield = rfields[cols_dim] if cols_dim else None
        col_repr = pt._represent_method(cols_dim)
        icols = pt.col
        cols = []
        for i in range(pt.numcols):
            icol = icols[i]
            header = {"value": icol.value,
                      "text": icol.text if "text" in icol
                                        else col_repr(icol.value),
                      }
            cols.append((i, icol[layer], header))
        pt._sortdim(cols, cols_rfield, index=2)

        return rows, cols

    # -------------------------------------------------------------------------
    def listrepr(self, cell, rfield, represent, fk=True):
        """
            Represent and sort a list of cell values (for "list" aggregation
            method)

            Args:
                cell - the cell data
                rfield - the fact S3ResourceField
                represent - representation method for the fact field
                fk - fact field is a foreign key

            Returns:
                sorted list of represented cell values
        """

        pt = self.pt
        records = pt.records

        colname = rfield.colname

        lookup = self.lookup
        valuemap = self.valuemap

        keys = []

        for record_id in cell["records"]:
            record = records[record_id]
            try:
                fvalue = record[colname]
            except AttributeError:
                continue

            if fvalue is None:
                continue
            if type(fvalue) is not list:
                fvalue = [fvalue]

            for v in fvalue:
                if v is None:
                    continue
                if fk:
                    if v not in keys:
                        keys.append(v)
                    if v not in lookup:
                        lookup[v] = represent(v)
                else:
                    if v not in valuemap:
                        next_id = len(valuemap)
                        valuemap[v] = next_id
                        keys.append(next_id)
                        lookup[next_id] = represent(v)
                    else:
                        prev_id = valuemap[v]
                        if prev_id not in keys:
                            keys.append(prev_id)

        keys.sort(key=lambda i: lookup[i])
        items = [s3_str(lookup[key]) for key in keys if key in lookup]

        return items

# =============================================================================
def dt_format_translate(pyfmt):
    """
        Translates a Python datetime format string into an
        Excel datetime format string

        Args:
            pyfmt: the Python format string

        Returns:
            the Excel datetime format string
    """

    translate = {"%a": "ddd",
                 "%A": "dddd",
                 "%b": "mmm",
                 "%B": "mmmm",
                 "%c": "",
                 "%d": "dd",
                 "%f": "",
                 "%H": "hh",
                 "%I": "hh",
                 "%j": "",
                 "%m": "mm",
                 "%M": "mm",
                 "%p": "AM/PM",
                 "%S": "ss",
                 "%U": "",
                 "%w": "",
                 "%W": "",
                 "%x": "",
                 "%X": "",
                 "%y": "yy",
                 "%Y": "yyyy",
                 "%z": "",
                 "%Z": "",
                 }

    PERCENT = "__percent__"
    xlfmt = str(pyfmt).replace("%%", PERCENT)

    for tag, translation in translate.items():
        xlfmt = xlfmt.replace(tag, translation)

    return xlfmt.replace(PERCENT, "%")

# =============================================================================
def to_int(string):
    """
        Convert a string representation of an integer back into an int
            - takes thousands-separator into account
            - strips any leading/trailing blanks

        Args:
            string: the string representation

        Returns:
            integer value

        Raises:
            ValueError if the string cannot be converted
    """

    sep = current.deployment_settings.get_L10n_thousands_separator()

    try:
        value = int(string.strip().replace(sep, ""))
    except (ValueError, TypeError, AttributeError):
        raise ValueError("not an integer number")
    return value

# =============================================================================
def to_float(string):
    """
        Convert a string representation of a float back into an float
            - takes thousands-/decimal-separators into account
            - strips any leading/trailing blanks

        Args:
            string: the string representation

        Returns:
            floating point value

        Raises:
            ValueError if the string cannot be converted
    """

    settings = current.deployment_settings
    tsep = settings.get_L10n_thousands_separator()
    dsep = settings.get_L10n_decimal_separator()

    try:
        value = float(string.strip().replace(tsep, "").replace(dsep, "."))
    except (ValueError, TypeError, AttributeError):
        raise ValueError("not a floating point number")
    return value

# END =========================================================================
