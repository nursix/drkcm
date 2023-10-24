"""
    GIMS Shelter Management Extension

    Copyright: 2022 (c) AHSS

    Permission is hereby granted, free of charge, to any person
    obtaining a copy of this software and associated documentation
    files (the "Software"), to deal in the Software without
    restriction, including without limitation the rights to use,
    copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following
    conditions:

    The above copyright notice and this permission notice shall be
    included in all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
    EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
    OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
    NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
    HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
    WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
    OTHER DEALINGS IN THE SOFTWARE.
"""

__all__ = ("CRReceptionCenterModel",
           )

import datetime
import json

from collections import OrderedDict

from gluon import current, Field, HTTP, URL, \
                  A, DIV, INPUT, OPTION, SELECT, SPAN, \
                  TABLE, TBODY, TD, TFOOT, TH, THEAD, TR, \
                  IS_EMAIL, IS_EMPTY_OR, IS_INT_IN_RANGE, IS_IN_SET, IS_NOT_EMPTY
from gluon.contenttype import contenttype
from gluon.storage import Storage

from core import CustomController, CRUDMethod, DataModel, DateField, FS, \
                 S3Report, S3Duplicate, LocationSelector, S3PriorityRepresent, \
                 S3Represent, FieldTemplate, S3SQLCustomForm, \
                 IS_ONE_OF, IS_PHONE_NUMBER_MULTI, \
                 get_form_record_id, CommentsField, \
                 s3_str, get_filter_options, \
                 LocationFilter, OptionsFilter, TextFilter

COMPUTE_ALLOCABLE_CAPACITY = False

# =============================================================================
class CRReceptionCenterModel(DataModel):
    """ Custom Model for Reception Centers """

    names = ("cr_reception_center_type",
             "cr_reception_center",
             "cr_reception_center_status",
             )

    def model(self):

        T = current.T
        db = current.db

        s3 = current.response.s3
        crud_strings = s3.crud_strings

        define_table = self.define_table
        configure = self.configure

        # Reusable field for population-type fields
        population = FieldTemplate("population", "integer",
                                   default = 0,
                                   requires = IS_INT_IN_RANGE(0),
                                   represent = lambda v, row=None: v if v != None else "-",
                                   )

        # ---------------------------------------------------------------------
        # Reception Center Types
        #
        tablename = "cr_reception_center_type"
        define_table(tablename,
                     Field("name",
                           label = T("Name"),
                           requires = IS_NOT_EMPTY(),
                           ),
                     CommentsField(),
                     )

        # Representation
        type_represent = S3Represent(lookup=tablename)

        # Table configuration
        configure(tablename,
                  deduplicate = S3Duplicate(),
                  )

        # CRUD Strings
        crud_strings[tablename] = Storage(
            label_create = T("Add Facility Type"),
            title_display = T("Type Details"),
            title_list = T("Reception Center Types"),
            title_update = T("Edit Facility Type"),
            label_list_button = T("List Facility Types"),
            label_delete_button = T("Delete Facility Type"),
            msg_record_created = T("Facility Type added"),
            msg_record_modified = T("Facility Type updated"),
            msg_record_deleted = T("Facility Type deleted"),
            msg_list_empty = T("No Facility Types currently defined"),
            )

        # ---------------------------------------------------------------------
        # Reception Centers
        #
        status_opts = (("OP", T("operating")),
                       ("SB", T("standby")),
                       ("NA", T("closed")),
                       )
        status = FieldTemplate("status",
                               default = "OP",
                               requires = IS_IN_SET(status_opts, zero=None),
                               represent = S3PriorityRepresent(dict(status_opts),
                                                               {"OP": "green",
                                                                "SB": "amber",
                                                                "NA": "grey"
                                                                }).represent,
                               )

        tablename = "cr_reception_center"
        define_table(tablename,
                     Field("name",
                           label = T("Name"),
                           requires = IS_NOT_EMPTY(),
                           ),
                     self.org_organisation_id(
                        comment = None,
                        requires = self.org_organisation_requires(required = True,
                                                                  updateable = True,
                                                                  ),
                        ),
                     self.gis_location_id(
                        widget = LocationSelector(levels = ("L1", "L2", "L3", "L4"),
                                                  required_levels = ("L1", "L2", "L3"),
                                                  show_address = True,
                                                  show_postcode = True,
                                                  show_map = True,
                                                  ),
                        ),
                     Field("type_id", "reference cr_reception_center_type",
                           label = T("Facility Type"),
                           requires = IS_ONE_OF(db, "cr_reception_center_type.id",
                                                type_represent,
                                                ),
                           represent = type_represent,
                           ),
                     status(),

                     Field("contact_name",
                           label = T("Contact Name"),
                           represent = lambda v, row=None: v if v else "-",
                           ),
                     Field("phone",
                           label = T("Phone"),
                           requires = IS_EMPTY_OR(IS_PHONE_NUMBER_MULTI()),
                           represent = lambda v, row=None: v if v else "-",
                           ),
                     Field("email",
                           label = T("Email"),
                           requires = IS_EMPTY_OR(IS_EMAIL()),
                           represent = lambda v, row=None: v if v else "-",
                           ),

                     # Capacity
                     population("capacity",
                                label = T("Capacity"),
                                comment = T("The maximum (total) capacity as number of people"),
                                ),
                     population("allocable_capacity_estimate",
                                # This is not a free estimate, but rather calculated
                                # by quota - hence the label should say "calculated"
                                # rather than "estimated":
                                label = T("Calculated Allocable Capacity"),
                                comment = T("Calculated total number of allocable places (by quota)"),
                                ),
                     population("allocable_capacity",
                                label = T("Allocable Capacity"),
                                comment = T("The total number of allocable places"),
                                # Computed onaccept?
                                readable = not COMPUTE_ALLOCABLE_CAPACITY,
                                writable = not COMPUTE_ALLOCABLE_CAPACITY,
                                ),

                     # Population
                     population("population_registered",
                                label = T("Current Population (Registered)"),
                                ),
                     population("population_unregistered",
                                label = T("Current Population (Unregistered)"),
                                ),
                     population(label = T("Current Population (Total)"),
                                # Computed onaccept
                                readable = False,
                                writable = False,
                                ),

                     # Free Capacitiy
                     population("free_capacity",
                                label = T("Free Capacity"),
                                comment = T("The number of free places"),
                                # Computed onaccept
                                readable = False,
                                writable = False,
                                ),
                     population("free_allocable_capacity",
                                label = T("Free Allocable Capacity"),
                                comment = T("The number of free allocable places"),
                                # Computed onaccept?
                                readable = COMPUTE_ALLOCABLE_CAPACITY,
                                writable = COMPUTE_ALLOCABLE_CAPACITY,
                                ),

                     # Utilization Rates
                     Field("utilization_rate", "integer",
                           label = T("Utilization %"),
                           represent = self.occupancy_represent,
                           readable = False,
                           writable = False,
                           ),
                     Field("occupancy_rate", "integer",
                           label = T("Occupancy %"),
                           represent = self.occupancy_represent,
                           readable = False,
                           writable = False,
                           ),

                     # TODO deprecate:
                     population("available_capacity", # => free_capacity
                                readable = False,
                                writable = False,
                                ),
                     population("allocatable_capacity", # => free_allocable_capacity
                                readable = False,
                                writable = False,
                                ),
                     Field("occupancy", "integer", # => utilization_rate
                           readable = False,
                           writable = False,
                           ),

                     CommentsField(),
                     )

        # Components
        self.add_components(tablename,
                            cr_reception_center_status = {"name": "status",
                                                          "joinby": "facility_id",
                                                          },
                            )

        # Representation
        facility_represent = S3Represent(lookup=tablename)

        # List fields
        list_fields = ["name",
                       "type_id",
                       "location_id",
                       "status",
                       "capacity",
                       "population_registered",
                       "population_unregistered",
                       "free_capacity",
                       "free_allocable_capacity",
                       "occupancy",
                       ]

        # Filter widgets
        filter_widgets = [TextFilter(["name",
                                      ],
                                     label = T("Search"),
                                     ),
                          OptionsFilter("type_id",
                                        options = get_filter_options("cr_reception_center_type"),
                                        ),
                          OptionsFilter("status",
                                        default = "OP",
                                        options = OrderedDict(status_opts),
                                        cols = 3,
                                        sort = False,
                                        ),
                          ]

        if COMPUTE_ALLOCABLE_CAPACITY:
            allocable_capacity = None
            free_allocable_capacity = "free_allocable_capacity"
        else:
            allocable_capacity = "allocable_capacity"
            free_allocable_capacity = None

        crud_form = S3SQLCustomForm(# ---- Facility ----
                                    "organisation_id",
                                    "name",
                                    "type_id",
                                    "status",
                                    # ---- Address ----
                                    "location_id",
                                    # ---- Capacity & Population ----
                                    "capacity",
                                    "allocable_capacity_estimate",
                                    allocable_capacity,
                                    "population_registered",
                                    "population_unregistered",
                                    free_allocable_capacity,
                                    # ---- Contact Information ----
                                    "contact_name",
                                    "phone",
                                    "email",
                                    # ---- Comments ----
                                    "comments",
                                    )

        subheadings = {"organisation_id": T("Facility"),
                       "location_id": T("Address"),
                       "capacity": T("Capacity / Occupancy"),
                       "contact_name": T("Contact Information"),
                       "comments": T("Comments"),
                       }

        # Report options
        axes = ["location_id$L3",
                "location_id$L2",
                "location_id$L1",
                "organisation_id",
                "type_id",
                "status",
                ]

        report_options = {
            "rows": axes,
            "cols": axes,
            "fact": [(T("Number of Facilities"), "count(id)"),
                     (T("Current Population (Total)"), "sum(population)"),
                     (T("Current Population (Registered)"), "sum(population_registered)"),
                     (T("Current Population (Unregistered)"), "sum(population_unregistered)"),
                     (T("Utilization % (Average)"), "avg(utilization_rate)"),
                     (T("Occupancy % (Average)"), "avg(occupancy_rate)"),
                     (T("Maximum Capacity"), "sum(capacity)"),
                     (T("Allocable Capacity"), "sum(allocable_capacity)"),
                     (T("Free Capacity"), "sum(free_capacity)"),
                     (T("Free Allocable Capacity"), "sum(free_allocable_capacity)"),
                     ],
            "defaults": {"rows": "location_id$L2",
                         "cols": "type_id",
                         "fact": "sum(population)",
                         "totals": True,
                         },
            }

        # Table configuration
        configure(tablename,
                  crud_form = crud_form,
                  subheadings = subheadings,
                  deduplicate = S3Duplicate(primary = ("name",),
                                            secondary = ("organisation_id",),
                                            ),
                  filter_widgets = filter_widgets,
                  list_fields = list_fields,
                  report_options = report_options,
                  onaccept = self.reception_center_onaccept,
                  )

        # Overview method
        self.set_method(tablename,
                        method = "overview",
                        action = CapacityOverview,
                        )
        self.set_method(tablename,
                        method = "occupancy",
                        action = OccupancyData,
                        )

        # CRUD Strings
        crud_strings[tablename] = Storage(
            label_create = T("Add Facility"),
            title_display = T("Facility Details"),
            title_list = T("Reception Centers"),
            title_update = T("Edit Facility"),
            label_list_button = T("List Facilities"),
            label_delete_button = T("Delete Facility"),
            msg_record_created = T("Facility added"),
            msg_record_modified = T("Facility updated"),
            msg_record_deleted = T("Facility deleted"),
            msg_list_empty = T("No Facilities currently registered"),
            )

        # ---------------------------------------------------------------------
        # Reception Center Status
        #
        tablename = "cr_reception_center_status"
        define_table(tablename,
                     Field("facility_id", "reference cr_reception_center",
                           label = T("Facility"),
                           ondelete = "CASCADE",
                           requires = IS_ONE_OF(db, "cr_reception_center.id",
                                                facility_represent,
                                                ),
                           represent = facility_represent,
                           ),
                     DateField(default = "now",
                               ),
                     DateField("date_until",
                               readable = False,
                               writable = False,
                               ),
                     status(),

                     population("capacity",
                                label = T("Capacity"),
                                comment = T("The maximum (total) capacity as number of people"),
                                ),
                     population("allocable_capacity_estimate",
                                label = T("Calculated Allocable Capacity"),
                                comment = T("Calculated total number of allocable places (by quota)"),
                                ),
                     population("allocable_capacity",
                                label = T("Allocable Capacity"),
                                comment = T("The total number of allocable places"),
                                ),

                     population("population_registered",
                                label = T("Current Population (Registered)"),
                                ),
                     population("population_unregistered",
                                label = T("Current Population (Unregistered)"),
                                ),
                     population(label = T("Current Population (Total)"),
                                ),

                     population("free_capacity",
                                label = T("Free Capacity"),
                                comment = T("The number of free places"),
                                ),
                     population("free_allocable_capacity",
                                label = T("Free Allocable Capacity"),
                                comment = T("The number of free allocable places"),
                                ),
                     Field("utilization_rate", "integer",
                           label = T("Utilization %"),
                           represent = self.occupancy_represent,
                           ),
                     Field("occupancy_rate", "integer",
                           label = T("Occupancy %"),
                           represent = self.occupancy_represent,
                           ),

                     # TODO deprecate:
                     population("available_capacity", # => free_capacity
                                readable = False,
                                writable = False,
                                ),
                     population("allocatable_capacity", # => free_allocable_capacity
                                readable = False,
                                writable = False,
                                ),
                     Field("occupancy", "integer", # => utilization_rate
                           readable = False,
                           writable = False,
                           ),
                     )

        # Filter Widgets
        filter_widgets = [TextFilter(["facility_id$name",
                                      ],
                                     label = T("Search"),
                                     ),
                          OptionsFilter("facility_id$type_id",
                                        options = get_filter_options("cr_reception_center_type"),
                                        ),
                          OptionsFilter("status",
                                        options = OrderedDict(status_opts),
                                        cols = 3,
                                        default = ["OP", "SB"],
                                        sort = False,
                                        ),
                          OptionsFilter("facility_id",
                                        options = get_filter_options("cr_reception_center"),
                                        hidden = True,
                                        ),
                          LocationFilter("facility_id$location_id",
                                         levels = ["L2", "L3"],
                                         hidden = True,
                                         ),
                          ]

        # Timeplot options
        facts = [(T("Current Population (Total)"), "sum(population)"),
                 (T("Current Population (Registered)"), "sum(population_registered)"),
                 (T("Current Population (Unregistered)"), "sum(population_unregistered)"),
                 (T("Utilization % (Average)"), "avg(utilization_rate)"),
                 (T("Occupancy % (Average)"), "avg(occupancy_rate)"),
                 (T("Maximum Capacity"), "sum(capacity)"),
                 (T("Allocable Capacity"), "sum(allocable_capacity)"),
                 (T("Free Capacity"), "sum(free_capacity)"),
                 (T("Free Allocable Capacity"), "sum(free_allocable_capacity)"),
                 ]
        timeplot_options = {
            "facts": facts,
            "timestamp": ((T("per interval"), "date,date_until"),
                          #(T("cumulative"), "date"),
                          ),
            "time": ((T("Last 3 Months"), "-3 months", "", "days"),
                     (T("This Month"), "<-0 months", "", "days"),
                     (T("This Week"), "<-0 weeks", "", "days"),
                     ),
            "defaults": {"fact": facts[0],
                         "timestamp": "date,date_until",
                         "time": "<-0 months||days",
                         },
            }

        # Table configuration
        configure(tablename,
                  filter_widgets = filter_widgets,
                  insertable = False,
                  editable = False,
                  deletable = False,
                  orderby = "%s.date desc" % tablename,
                  timeplot_options = timeplot_options,
                  )

        # CRUD Strings
        crud_strings[tablename] = Storage(
            title_display = T("Status"),
            title_list = T("Status History"),
            label_list_button = T("Status History"),
            msg_list_empty = T("No Status History currently registered"),
            )

        # ---------------------------------------------------------------------
        # Pass names back to global scope (s3.*)
        #
        return None

    # -------------------------------------------------------------------------
    @classmethod
    def reception_center_onaccept(cls, form):
        """
            Onaccept of reception center
                - update total population
                - update free capacity
                - compute allocable / free allocable capacity
                - update utilization and occupancy rates
                - update status history
        """

        record_id = get_form_record_id(form)
        if not record_id:
            return

        table = current.s3db.cr_reception_center
        query = (table.id == record_id) & \
                (table.deleted == False)
        record = current.db(query).select(table.id,
                                          table.population_registered,
                                          table.population_unregistered,
                                          table.capacity,
                                          table.allocable_capacity_estimate,
                                          table.allocable_capacity,
                                          table.free_capacity,
                                          table.free_allocable_capacity,
                                          limitby = (0, 1),
                                          ).first()

        # Compute total population
        num_r = record.population_registered
        num_u = record.population_unregistered
        total = (num_r if num_r else 0) + (num_u if num_u else 0)
        update = {"population": total}

        # Compute free capacity
        capacity = record.capacity
        free_capacity = max(0, capacity - total) if capacity else 0
        update["free_capacity"] = free_capacity

        # Sanitize allocable_capacity_estimate
        estimate = record.allocable_capacity_estimate
        estimate = update["allocable_capacity_estimate"] = max(0, min(capacity, estimate))

        # Compute allocable capacity
        if COMPUTE_ALLOCABLE_CAPACITY:
            # Sanitize free_allocable_capacity
            free_allocable_capacity = record.free_allocable_capacity
            if free_allocable_capacity > free_capacity:
                free_allocable_capacity = free_capacity
            if free_allocable_capacity < 0:
                free_allocable_capacity = 0
            # Compute maximum allocable capacity
            allocable_capacity = min(capacity, free_allocable_capacity + total)
        else:
            # Sanitize allocable_capacity
            allocable_capacity = record.allocable_capacity
            if allocable_capacity > capacity:
                allocable_capacity = capacity
            if allocable_capacity < 0:
                allocable_capacity = 0
            # Compute free allocable capacity
            free_allocable_capacity = max(0, allocable_capacity - total)

        # Warn if actual allocable capacity is below estimate:
        # => this means that the policy makers work with an unrealistic
        #    estimate (i.e., too high), so the coordinating authority may
        #    need to report and/or correct this discrepancy (small, transient
        #    differences do not necessarily require action, however)
        # => correcting actions could involve:
        #    - optimizing capacity utilization (re-allocating places)
        #    - correcting the (maximum) capacity to actual numbers
        #    - adjusting the quota for estimate calculation
        if allocable_capacity < estimate:
            current.response.warning = current.T("Actual allocable capacity is lower than the calculated allocable capacity!")

        update["allocable_capacity"] = allocable_capacity
        update["free_allocable_capacity"] = free_allocable_capacity

        # Compute utilization rate (by maximum capacity)
        if capacity > 0:
            utilization_rate = total * 100 // capacity
        else:
            utilization_rate = 100
        update["utilization_rate"] = utilization_rate

        # Compute occupancy rate (by allocable capacity)
        if allocable_capacity > 0:
            occupancy_rate = total * 100 // allocable_capacity
        else:
            occupancy_rate = 100
        update["occupancy_rate"] = occupancy_rate

        record.update_record(**update)

        # Update the status history
        cls.update_status_history(record.id)

    # -------------------------------------------------------------------------
    @staticmethod
    def update_status_history(facility_id):
        """
            Updates the status history of a facility

            Args:
                facility_id: the cr_reception_center record ID
        """

        db = current.db
        s3db = current.s3db

        ftable = s3db.cr_reception_center
        stable = s3db.cr_reception_center_status

        status_fields = ("status",
                         "capacity",
                         "allocable_capacity_estimate",
                         "allocable_capacity",
                         "population",
                         "population_registered",
                         "population_unregistered",
                         "free_capacity",
                         "free_allocable_capacity",
                         "utilization_rate",
                         "occupancy_rate",
                         )

        # Get the reception center record
        fields = [ftable.id] + [ftable[fn] for fn in status_fields]
        query = (ftable.id == facility_id) & (ftable.deleted == False)
        facility = db(query).select(*fields, limitby=(0, 1)).first()

        # Look up the status record for today
        today = current.request.utcnow.date()
        query = (stable.facility_id == facility_id) & \
                (stable.date == today) & \
                (stable.deleted == False)
        status = db(query).select(stable.id, limitby = (0, 1)).first()
        if not status:
            # Create it
            data = {fn: facility[fn] for fn in status_fields}
            data["facility_id"] = facility.id
            status_id = data["id"] = stable.insert(**data)
            s3db.update_super(stable, status)
            current.auth.s3_set_record_owner(stable, status_id)
            s3db.onaccept(stable, status, method="create")
        else:
            # Update it
            update = {fn: facility[fn] for fn in status_fields}
            status.update_record(**update)
            s3db.onaccept(stable, status, method="update")

        # Update the previous record (set end-date)
        query = (stable.facility_id == facility_id) & \
                (stable.date < today) & \
                (stable.deleted == False)
        status = db(query).select(stable.id,
                                  orderby = ~stable.date,
                                  limitby = (0, 1),
                                  ).first()
        if status:
            status.update_record(date_until = today-datetime.timedelta(days=1))

    # -------------------------------------------------------------------------
    @staticmethod
    def occupancy_represent(value, row=None):
        """
            Representation of utilization/occupancy rate as decision aid,
            progress-bar style

            Args:
                value: the occupancy value (percentage, integer 0..>100)

            Returns:
                stylable DIV
        """

        if not value:
            value = 0
            css_class = "occupancy-0"
        else:
            reprval = value // 10 * 10 + 10
            if reprval > 100:
                css_class = "occupancy-exc"
            else:
                css_class = "occupancy-%s" % reprval

        return DIV("%s%%" % value,
                   DIV(_class="occupancy %s" % css_class),
                   _class="occupancy-bar",
                   )

# =============================================================================
class CapacityOverview(CRUDMethod):
    """
        Custom overview method for reception center capacities/populations
    """

    def apply_method(self, r, **attr):
        """
            Main entry point for CRUD controller

            Args:
                r: the CRUDRequest
                attr: controller parameters
        """

        if r.http == "GET":
            if r.representation == "html":
                output = self.overview(r, **attr)
            else:
                r.error(415, current.ERROR.BAD_FORMAT)
        else:
            r.error(405, current.ERROR.BAD_METHOD)

        return output

    # -------------------------------------------------------------------------
    def overview(self, r, **attr):
        """
            Builds elements for the capacity overview page

            Args:
                r: the CRUDRequest
                attr: controller parameters

            Notes:
                - uses templates/GIMS/views/capacity.html as view template
                - injects the necessary client-side scripts
        """

        T = current.T

        # Build components
        output = {"title": T("Reception Centers Overview"),
                  "table": self.render_table(),
                  "chart": self.render_chart(),
                  "export": self.export_widget(),
                  }

        # Inject JS
        self.inject_scripts()

        # Set view
        CustomController._view("GIMS", "capacity.html")

        return output

    # -------------------------------------------------------------------------
    def render_table(self):
        """
            Builds a TABLE with current capacity/occupancy data of all
            available reception centers

            Returns:
                TABLE
        """

        T = current.T
        auth = current.auth

        resource = self.resource

        # Include only facilities in operation or standby
        resource.add_filter(FS("status").belongs(("OP", "SB")))

        # Fields to show (in order)
        if auth.s3_has_roles(("AFA_COORDINATOR", "AFA_MANAGER")):
            list_fields = [(T("Place"), "location_id$L3"),
                           (T("Facility"), "name"),
                           "status",
                           "capacity",
                           #"allocable_capacity_estimate",
                           "allocable_capacity",
                           "population",
                           "population_unregistered",
                           "free_capacity",
                           "comments",
                           "occupancy_rate",
                           "utilization_rate",
                           ]
            if auth.s3_has_role("AFA_COORDINATOR"):
                list_fields.insert(4, "allocable_capacity_estimate")
        else:
            list_fields = [(T("Place"), "location_id$L3"),
                           (T("Facility"), "name"),
                           "status",
                           (T("Capacity"), "allocable_capacity_estimate"),
                           "population",
                           "population_unregistered",
                           (T("Free Capacity"), "free_capacity_estimate"), # Custom
                           "comments",
                           (T("Occupancy %"), "occupancy_rate_estimate"), # Custom
                           ]

        # Extract the data (least occupied facilities first)
        data = resource.select(list_fields + ["id", "location_id$L4"],
                               represent = True,
                               raw_data = True,
                               limit = None,
                               orderby = ["gis_location.L3 asc",
                                          "gis_location.L4 asc",
                                          "cr_reception_center.capacity desc",
                                          "cr_reception_center.occupancy desc",
                                          ],
                               )
        rfields = data.rfields

        # Display columns
        exclude = ("gis_location.L4")
        columns = [(rfield.colname, rfield.label)
                   for rfield in rfields
                   if rfield.show and rfield.ftype != "id" and rfield.colname not in exclude
                   ]

        # Label row
        thead = THEAD(TR([TH(col[1]) for col in columns]))

        # Data rows
        capacity_fields = ["capacity",
                           "allocable_capacity_estimate",
                           "allocable_capacity",
                           "population",
                           "population_unregistered",
                           "free_capacity_estimate",
                           "free_capacity",
                           "free_allocable_capacity",
                           ]
        totals = {fn: 0 for fn in capacity_fields}

        tbody = TBODY()
        append = tbody.append
        for row in data.rows:
            # Add to totals
            raw = row._row
            for fn in capacity_fields:
                value = raw["cr_reception_center.%s" % fn]
                if value:
                    totals[fn] += value

            # Render name as link to the facility
            name = row["cr_reception_center.name"]
            record_id = raw["cr_reception_center.id"]
            if name and record_id:
                url = URL(c="cr", f="reception_center", args=[record_id])
                row["cr_reception_center.name"] = A(name, _href=url)

            # Append data row to table
            if raw["gis_location.L4"]:
                # Replace L3 by L4
                display = [col[0] if col[0] != "gis_location.L3" else "gis_location.L4" for col in columns]
            else:
                display = [col[0] for col in columns]
            append(TR([TD(row[colname]) for colname in display]))

        # Compute total utilization/occupancy rates
        population = totals["population"]
        rates = {"capacity": "utilization_rate",
                 "allocable_capacity": "occupancy_rate",
                 "allocable_capacity_estimate": "occupancy_rate_estimate",
                 }
        represent = CRReceptionCenterModel.occupancy_represent
        for fname, rname in rates.items():
            capacity = totals[fname]
            if capacity > 0:
                rate = population * 100 // capacity
            else:
                rate = 100
            totals[rname] = represent(rate)

        # Footer with totals
        tr = TR()
        append = tr.append
        for rfield in rfields:
            # Skip hidden and excluded fields
            if not rfield.show or rfield.ftype == "id" or rfield.colname in exclude:
                continue
            fn = rfield.fname
            if fn in totals:
                append(TD(totals[fn]))
            elif fn == "L3":
                append(TD(T("Total##set")))
            else:
                append(TD())
        tfoot = TFOOT(tr)

        # Combine table
        table = TABLE(thead, tbody, tfoot)

        return table

    # -------------------------------------------------------------------------
    def render_chart(self):
        """
            Builds the HTML structure for the population chart, with
            time series data injected

            Returns:
                DIV
        """

        DAYS = 365

        T = current.T

        db = current.db
        s3db = current.s3db

        # Look up reception centers
        ftable = s3db.cr_reception_center
        query = current.auth.s3_accessible_query("read", ftable) & \
                (ftable.deleted == False)
        facilities = db(query).select(ftable.id, ftable.name)

        # Use facility names as labels for data series
        labels = {facility.id: facility.name for facility in facilities}
        labels[0] = s3_str(T("Total##set"))

        # Start time
        start = datetime.datetime.utcnow().date() - datetime.timedelta(days=DAYS)
        timestmp = int(datetime.datetime.combine(start, datetime.time(12,0,0)).timestamp())

        # Extract and format the data
        data = {"type": s3_str(T("Population##shelter")),
                "population": self.get_time_series(facilities, DAYS),
                "labels": labels,
                "start": timestmp,
                }

        # Compose HTML
        chart = DIV(INPUT(_id = "history-data",
                          _type = "hidden",
                          _value = json.dumps(data,
                                              separators = (",", ":"),
                                              ensure_ascii = False,
                                              ),
                          ),
                    DIV(_id="history-chart"),
                    _class = "capacity-chart",
                    )

        return chart

    # -------------------------------------------------------------------------
    @staticmethod
    def get_time_series(facilities, days):
        """
            Produces a time series of population numbers

            Args:
                facilities: the reception center Rows (must include record IDs)
                days: number of days before today

            Returns:
                a JSON-serializable dict {facility_id: [population_number, ...]}

            Note:
                - result dict includes totals with facility_id=0
        """

        db = current.db
        s3db = current.s3db

        stable = s3db.cr_reception_center_status

        start = datetime.datetime.utcnow().date() - datetime.timedelta(days=days)

        # Lookup initial population numbers (per facility)
        population = {}
        for facility in facilities:
            facility_id = facility.id
            query = (stable.facility_id == facility_id) & \
                    (stable.date < start) & \
                    (stable.deleted == False)
            initial = db(query).select(stable.status,
                                       stable.population,
                                       orderby = ~stable.date,
                                       limitby = (0, 1),
                                       ).first()

            if initial and initial.population and initial.status in ("OP", "SB"):
                population[facility_id] = [initial.population] * days
            else:
                population[facility_id] = [0] * days

        # Lookup all subsequent status entries and fill the matrix
        query = (stable.facility_id.belongs(population.keys())) & \
                (stable.date != None) & \
                (stable.date >= start) & \
                (stable.deleted == False)
        rows = db(query).select(stable.facility_id,
                                stable.date,
                                stable.status,
                                stable.population,
                                orderby = stable.date,
                                )
        for row in rows:
            facility_id = row.facility_id

            day = (row.date - start).days

            series = population[facility_id]
            series[(day-1):days] = [row.population if row.population else 0] * (days-day+1)

        # Compute totals
        total = [sum(series[i] for series in population.values()) for i in range(days)]
        population[0] = total

        return population

    # -------------------------------------------------------------------------
    @staticmethod
    def export_widget():
        """
            Widget to download occupancy raw data

            Returns:
                a DIV
        """

        T = current.T

        # Get the date of the first status record
        table = current.s3db.cr_reception_center_status
        query = (table.status.belongs(("OP", "SB"))) & \
                (table.population != None) & \
                (table.date != None) & \
                (table.deleted == False)
        row = current.db(query).select(table.date,
                                       limitby = (0, 1),
                                       orderby = table.date,
                                       ).first()

        # Determine year range
        current_year = current.request.utcnow.year
        first_year = row.date.year if row else current_year

        # Build the year selector
        selector = SELECT(_id="data-year")
        for year in range(first_year, current_year+1):
            year_str = str(year)
            option = OPTION(year_str,
                            _value = year_str,
                            _selected = "selected" if year == current_year else None,
                            )
            selector.append(option)

        # Build the widget
        return DIV(SPAN(T("Occupancy Data"),
                        _class = "action-lbl",
                        ),
                   selector,
                   A(T("download##verb"),
                     _id = "data-download",
                     _class = "action-lnk",
                     data = {"url": URL(c = "cr",
                                        f = "reception_center",
                                        args = ["occupancy.xlsx"],
                                        ),
                             },
                     ),
                   _class="occupancy-export",
                   )

    # -------------------------------------------------------------------------
    @staticmethod
    def inject_scripts():
        """
            Inject client-side scripts for capacity overview page
        """

        s3 = current.response.s3

        S3Report.inject_d3() # D3+NVD3

        script = "/%s/static/themes/RLP/js/capacity.js" % current.request.application
        if script not in s3.scripts:
            s3.scripts.append(script)

# =============================================================================
class OccupancyData(CRUDMethod):
    """ Method to produce a occupancy data sheet """

    # -------------------------------------------------------------------------
    def apply_method(self, r, **attr):
        """
            Entry point for CRUD controller

            Args:
                r: the CRUDRequest
                attr: controller parameters
            Returns:
                output data for view
        """

        if r.http == "GET":
            if r.representation == "xlsx":
                output = self.export_data(r, **attr)
            else:
                r.error(415, current.ERROR.BAD_FORMAT)
        else:
            r.error(405, current.ERROR.BAD_METHOD)

        return output

    # -------------------------------------------------------------------------
    def export_data(self, r, **attr):
        """
            Exports the occupancy data in Excel format

            Args:
                r: the CRUDRequest
                attr: controller parameters
            Returns:
                output data for view
        """

        year = r.get_vars.get("year")
        if year:
            try:
                year = int(year)
            except ValueError:
                year = None
            if 2100 < year < 2000:
                year = None
        if not year:
            year = datetime.datetime.utcnow().year

        facilities = self.get_facilities()

        matrix = self.get_occupancy(facilities, year)#, start, end)

        response = current.response

        title = "Belegungszahlen %04d" % year

        # Set response headers
        filename = "%s.xlsx" % title
        disposition = "attachment; filename=\"%s\"" % filename
        response = current.response
        response.headers["Content-Type"] = contenttype(".xlsx")
        response.headers["Content-disposition"] = disposition

        return self.write_xlsx(facilities, year, matrix)

    # -------------------------------------------------------------------------
    @staticmethod
    def get_facilities():
        """
            Selects the relevant (accessible) facilities

            Returns:
                Rows
        """

        s3db = current.s3db

        ftable = s3db.cr_reception_center
        gtable = s3db.gis_location

        left = gtable.on(gtable.id == ftable.location_id)

        query = current.auth.s3_accessible_query("read", ftable) & \
                (ftable.deleted == False)

        rows = current.db(query).select(ftable.id,
                                        ftable.name,
                                        gtable.L3,
                                        gtable.L4,
                                        orderby = [gtable.L2,
                                                   gtable.L3,
                                                   gtable.L4,
                                                   ftable.name,
                                                   ],
                                        left = left,
                                        )
        facilities = []
        for row in rows:
            facility = row.cr_reception_center
            location = row.gis_location
            facility.place = location.L4 if location.L4 else location.L3
            facilities.append(facility)

        return facilities

    # -------------------------------------------------------------------------
    @classmethod
    def get_occupancy(cls, facilities, year):
        """
            Returns the occupancy data for the facilities for a certain year

            Args:
                facilities: the facility Rows
                year: the year

            Returns:
                Array [[date, value, value, ...], ...] with the
                occupancy numbers for each facility and each day of the year
        """

        db = current.db
        s3db = current.s3db

        start = datetime.date(year, 1 ,1)
        end = datetime.date(year, 12, 31)

        # Lookup initial occupancy numbers
        initial = cls.get_initial_occupancy(facilities, start)

        # Lookup occupancy updates within the year
        stable = s3db.cr_reception_center_status
        query = (stable.facility_id.belongs(initial.keys())) & \
                (stable.date != None) & \
                (stable.date >= start) & \
                (stable.date <= end) & \
                (stable.deleted == False)
        rows = db(query).select(stable.facility_id,
                                stable.date,
                                stable.status,
                                stable.population,
                                orderby = stable.date,
                                )

        # Order the updates by date
        data = {}
        for row in rows:
            facility_id = row.facility_id
            if facility_id in data:
                series = data[facility_id]
            else:
                series = data[facility_id] = {}
            series[row.date] = row.population if row.status in ("OP", "SB") else -1

        # Convert into a continuous value matrix
        date = start
        today = datetime.datetime.utcnow().date()
        matrix = []
        while date <= end:

            row, total = [date], 0
            for facility in facilities:
                facility_id = facility.id

                series = data.get(facility_id)
                value = series.get(date) if series else None

                if date > today or value == -1:
                    initial[facility_id] = value = None
                elif value is None:
                    value = initial.get(facility_id)
                else:
                    initial[facility_id] = value

                row.append(value)
                if value:
                    total += value

            if any(value != None for value in row[1:]):
                row.append(total)
            else:
                row.append(None)
            matrix.append(row)
            date += datetime.timedelta(days=1)

        return matrix

    # -------------------------------------------------------------------------
    @classmethod
    def write_xlsx(cls, facilities, year, matrix):
        """
            Produces an Excel document from the occupancy data matrix

            Args:
                facilities: the facility Rows
                year: the year
                matrix: the data matrix from get_occupancy
        """

        # Import OpenPyXL
        try:
            from openpyxl import Workbook
        except ImportError:
            error = current.T("Export failed: OpenPyXL library not installed on server")
            current.log.error(error)
            raise HTTP(503, body=error)

        # Create the workbook
        wb = Workbook(iso_dates=True)

        # Add named styles
        from core import XLSXWriter
        XLSXWriter.add_styles(wb, use_color=False, even_odd=True)

        # Write annual data (first sheet)
        cls.add_annual_data(wb, facilities, matrix, year)

        # Write monthly data (subsequent sheets)
        cls.add_monthly_data(wb, facilities, matrix)

        # Save workbook and read its contents
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            output = tmp.read()

        return output

    # -------------------------------------------------------------------------
    @classmethod
    def add_annual_data(cls, wb, facilities, matrix, year):
        """
            Writes annual statistics to the Excel workbook

            Args:
                wb: the workbook
                facilities: the facility Rows
                matrix: the occupancy data matrix
                year: the year in question
        """

        T = current.T

        # Use active data sheet
        ws = wb.active
        ws.title = "%04d" % year

        # Add label row
        labels = [T("Place"),
                  T("Facility"),
                  T("Days"),
                  T("Minimum"),
                  T("Average"),
                  T("Median"),
                  T("Q80"), # 80%-quantil
                  T("Maximum"),
                  ]
        label_row, column_widths = [], []
        for label in labels:
            label_str = s3_str(label)
            column_widths.append(len(label_str))
            label_row.append((label_str, label_str, None, "label"))
        cls.add_row(ws, label_row, column_widths)

        # Add data rows
        totals = cls.compute_totals(matrix)
        for i, (days, min_, avg_, med_, q80_, max_) in enumerate(totals):
            if i < len(facilities):
                facility = facilities[i]
                place = facility.place
                label = facility.name
                style = None
            else:
                place = ""
                label = s3_str(T("Total##set"))
                style = "label"
            cls.add_row(ws,
                        [(place, place, None, style),
                         (label, label, None, style),
                         (days, str(days), "0", style),
                         (min_, str(min_), "0", style),
                         (avg_, str(avg_), "0", style),
                         (med_, str(avg_), "0", style),
                         (q80_, str(q80_), "0", style),
                         (max_, str(max_), "0", style),
                         ],
                        column_widths,
                        )

        # Adjust column widths
        cls.adjust_column_widths(ws, column_widths)

        # Scroll only data rows, not labels
        ws.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    @classmethod
    def add_monthly_data(cls, wb, facilities, matrix):
        """
            Writes the monthly occupancy data to the Excel workbook

            Args:
                wb: the workbook
                facilities: the facility Rows
                matrix: the occupancy data matrix
        """

        T = current.T

        # Generate columns labels and initialize totals
        labels = [s3_str(T("Date"))]
        totals = []
        for facility in facilities:
            labels.append(s3_str(facility.name))
            totals.append((0, 0))
        labels.append(s3_str(T("Total##set")))
        totals.append((0, 0))
        label_widths = [len(label) + 1 for label in labels]

        AVG = s3_str(T("Average"))
        if label_widths[0] < len(AVG) + 1:
            label_widths[0] = len(AVG) + 1

        # Date format
        from core.formats.xlsx import dt_format_translate
        settings = current.deployment_settings
        date_format = dt_format_translate(settings.get_L10n_date_format())

        from core import S3DateTime
        date_represent = S3DateTime.date_represent

        # Write one sheet per month
        ws = None
        current_month = (None, None)

        for row in matrix:

            date = row[0]
            month = (date.year, date.month)

            last = (date + datetime.timedelta(days=1)).month != date.month

            if month != current_month:

                current_month = month
                values = [[] for _ in range(len(row) - 1)]
                column_widths = label_widths
                row_index = 0

                # Create new sheet
                ws = wb.create_sheet(title="%04d-%02d" % month)

                # Add column labels
                label_row = [(l, l, None, "label") for l in labels]
                cls.add_row(ws, label_row, column_widths)

                # Scroll only data rows, not labels
                ws.freeze_panes = "A2"

            row_index += 1
            row_style = "odd" if row_index % 2 else "even"

            col_idx = 0
            outrow = []

            # Write outrow
            for col_idx, value in enumerate(row):

                if col_idx == 0:
                    formatter = date_represent
                    number_format = date_format
                else:
                    formatter = s3_str
                    number_format = "0"

                if value is not None:
                    if col_idx > 0:
                        values[col_idx - 1].append(value)
                    formatted = formatter(value)
                else:
                    formatted = ""
                item = (value, formatted, number_format, row_style)

                outrow.append(item)

            # Append row to current sheet
            cls.add_row(ws, outrow, column_widths)

            # Post-process last row
            if last:
                # Compute and write averages
                outrow = [(AVG, AVG, None, "label")]

                for i, col_values in enumerate(values):
                    if len(col_values):
                        total, days = sum(col_values), len(col_values)
                        value = round(total / days, 0)
                    else:
                        total, days = 0, 0
                        value = 0
                    outrow.append((value, str(value), "0", "label"))

                    # Update totals
                    t = totals[i]
                    totals[i] = t[0] + total, t[1] + days

                cls.add_row(ws, outrow, column_widths)

                # Adjust column widths
                cls.adjust_column_widths(ws, column_widths)

        return totals

    # -------------------------------------------------------------------------
    @staticmethod
    def add_row(ws, items, column_widths):
        """
            Writes a row to a worksheet, and updates columns widths as needed

            Args:
                ws: the worksheet
                items: the items,
                       array [(raw_value, formatted, number_format, style)]
                column_widths: mutable array of column widths
        """

        from openpyxl.cell import Cell

        row = []
        for i, item in enumerate(items):

            value, text, number_format, style = item

            if value is None:
                cell = Cell(ws)
            else:
                cell = Cell(ws, value=value)

            if style:
                cell.style = style
            if number_format:
                cell.number_format = number_format

            row.append(cell)

            width = len(text)
            if style == "label":
                width += 1
            if width > column_widths[i]:
                column_widths[i] = width

        ws.append(row)

    # -------------------------------------------------------------------------
    @staticmethod
    def adjust_column_widths(ws, column_widths):
        """
            Adjusts the column widths of a worksheet

            Args:
                ws: the worksheet
                column_widths: array of column widths
        """

        from openpyxl.utils import get_column_letter

        for i in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(i+1)].width = column_widths[i] * 1.23

    # -------------------------------------------------------------------------
    @staticmethod
    def get_initial_occupancy(facilities, start):
        """
            Determines the initial occupancy for the given facilities
            at the given start date (...from the last status update
            before that start date, if one exists)

            Args:
                facilities: the facility Rows
                start: the start date

            Returns:
                a dict {facility_id: number|None}
        """

        db = current.db
        s3db = current.s3db

        stable = s3db.cr_reception_center_status

        # Lookup initial population numbers (per facility)
        initial = {}
        for facility in facilities:
            facility_id = facility.id
            query = (stable.facility_id == facility_id) & \
                    (stable.date < start) & \
                    (stable.deleted == False)
            row = db(query).select(stable.status,
                                   stable.population,
                                   orderby = ~stable.date,
                                   limitby = (0, 1),
                                   ).first()

            if row and row.status in ("OP", "SB"):
                initial[facility_id] = row.population
            else:
                initial[facility_id] = None

        return initial

    # -------------------------------------------------------------------------
    @classmethod
    def compute_totals(cls, matrix):
        """
            Computes totals for the data matrix

            Args:
                matrix: the occupancy data matrix

            Returns:
                Array [(days, min, avg, median, q80, max), ...] with one
                tuple per facility, and one additional tuple for totals

            Notes:
                days = the number of days the facility was in use during the year
                min = the minimum occupancy
                avg = the average occupancy
                median = the median occupancy (=50% of days were at or above this value)
                q80 = the 80. percentile (=80% of days were at or below this value)
                max = the maximum occupancy
        """

        values = None

        for row in matrix:
            if values is None:
                values = [[] for _ in range(len(row) - 1)]
            for i, value in enumerate(row[1:]):
                if value is not None:
                    values[i].append(value)

        import statistics

        totals = []
        for series in values:
            days = len(series)
            if not days:
                totals.append((0, 0, 0, 0, 0, 0))
            else:
                totals.append((days,
                               min(series),
                               round(sum(series) / days, 0),
                               statistics.median(series),
                               cls.quantile(series, 0.8),
                               max(series),
                               ))

        return totals

    # -------------------------------------------------------------------------
    @staticmethod
    def quantile(a, q, precision=0):
        """
            Determines a quantile for array a, such that q% of the values
            in a are less than or equal to the quantile

            Args:
                a: array of int|float
                q: the quantile 0..100
                precision: the rounding precision for interpolated values

            Returns:
                the quantile value (float)
        """

        if not a:
            raise RuntimeError("empty sample")
        if len(a) == 1:
            return a[0]

        s = sorted(a)
        n = len(s)
        p = (n - 1) * q + 1

        i = int(p)
        if float(i) == p:
            # Choose
            quantile = float(s[i-1])
        else:
            # Interpolate
            a = float(s[i-1])
            b = float(s[i])
            quantile = a + (b - a) * (p - i)

        return round(quantile, precision)

# END =========================================================================
